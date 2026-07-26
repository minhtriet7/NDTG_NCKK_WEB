"""Benchmark metrics verifier for BanknoteAI, GPT and Gemini workbooks.

The service keeps the download contract intact: ``BenchmarkMetricsService``
returns a BytesIO xlsx. It does not trust source correctness columns; every
field-level result is recalculated from ground-truth and prediction columns.
"""

from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from statistics import stdev
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:  # pragma: no cover - exercised in the real FastAPI app.
    from fastapi import HTTPException
except ModuleNotFoundError:  # pragma: no cover - local logic tests only.
    class HTTPException(Exception):
        def __init__(self, status_code: int, detail: str):
            self.status_code = status_code
            self.detail = detail
            super().__init__(detail)

from app.utils.benchmark_metrics_math import (
    classification_report_dict,
    mean,
    report_summary,
)
from app.utils.benchmark_normalization import (
    MISSING_SENTINEL,
    calculate_field_correctness,
    normalize_boolean,
    normalize_country,
    normalize_currency,
    normalize_denomination,
)


MODEL_BANKNOTEAI = "BanknoteAI"
MODEL_GPT = "GPT"
MODEL_GEMINI = "Gemini"

HE_THONG_SHEET_CANDIDATES = ("HeThong", "Sheet4", "experiment_runs")
GPT_GEMINI_SHEET_CANDIDATES = ("GPT_GEMINI", "GPT_Gemini", "GPT Gemini")
MANIFEST_REQUIRED_COLUMNS = {
    "dataset_id",
    "image_id",
    "file_name",
    "ground_truth_country",
    "ground_truth_currency",
    "ground_truth_denomination",
}

HE_THONG_REQUIRED_COLUMNS = {
    "ground_truth_country",
    "ground_truth_currency",
    "ground_truth_denomination",
    "predicted_country",
    "predicted_currency",
    "predicted_denomination",
}

GPT_REQUIRED_COLUMNS = {
    "model_name",
    "gt_country",
    "gt_currency",
    "gt_denomination",
    "pred_country",
    "pred_currency",
    "pred_denomination",
}

RAW_VERIFIED_COLUMNS = [
    "dataset_id_verified",
    "model_name_verified",
    "source_sheet",
    "source_row",
    "normalized_ground_truth_country",
    "normalized_predicted_country",
    "normalized_ground_truth_currency",
    "normalized_predicted_currency",
    "normalized_ground_truth_denomination",
    "normalized_predicted_denomination",
    "country_correct_original",
    "currency_correct_original",
    "denomination_correct_original",
    "field_correct_count_original",
    "field_score_pct_original",
    "exact_match_original",
    "country_correct_verified",
    "currency_correct_verified",
    "denomination_correct_verified",
    "field_correct_count_verified",
    "field_score_pct_verified",
    "exact_match_verified",
    "has_complete_prediction",
    "exact_match_mismatch",
    "country_correct_mismatch",
    "currency_correct_mismatch",
    "denomination_correct_mismatch",
    "field_correct_count_mismatch",
]

ISSUE_HEADERS = [
    "code",
    "severity",
    "category",
    "dataset_id",
    "image_id",
    "model",
    "run_no",
    "file_name",
    "source_sheet",
    "source_row",
    "message",
    "expected",
    "actual",
]

SUMMARY_HEADERS = [
    "Dataset",
    "Model",
    "Total Samples",
    "Complete Predictions",
    "Correct Samples",
    "Coverage",
    "Accuracy",
    "Precision",
    "Recall",
    "F1",
    "Consensus Rate",
    "Status",
]

DIMENSION_HEADERS = [
    "Dataset",
    "Model",
    "Dimension",
    "Accuracy",
    "Macro Precision",
    "Macro Recall",
    "Macro F1",
]

PER_RUN_HEADERS = [
    "Dataset",
    "Model",
    "Run",
    "Total Samples",
    "Coverage",
    "Accuracy",
    "Macro Precision",
    "Macro Recall",
    "Macro F1",
    "Consensus Rate",
]

NOTES_HEADERS = [
    "Severity",
    "Dataset",
    "Model",
    "Image ID",
    "Run",
    "Note Type",
    "Current Value",
    "Expected Value",
    "Description",
]

MISSING_CLASS_LABEL = "MISSING"


def _header_key(value: Any) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def _is_blank_row(row: Sequence[Any]) -> bool:
    return not any(value is not None and str(value).strip() for value in row)


def _load_workbook(file_bytes: bytes) -> Any:
    try:
        return openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Không thể mở file Excel: {exc}",
        ) from exc


def _sheet_headers(ws: Any) -> List[str]:
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        return []
    return [_header_key(cell) for cell in first_row]


def _find_sheet(
    wb: Any,
    *,
    required_columns: Iterable[str],
    candidates: Sequence[str],
) -> Optional[str]:
    required = set(required_columns)
    for candidate in candidates:
        if candidate in wb.sheetnames:
            headers = set(_sheet_headers(wb[candidate]))
            if required.issubset(headers):
                return candidate
    for sheet_name in wb.sheetnames:
        headers = set(_sheet_headers(wb[sheet_name]))
        if required.issubset(headers):
            return sheet_name
    return None


def _read_sheet_rows(wb: Any, sheet_name: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    ws = wb[sheet_name]
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return [], []
    headers = [_header_key(cell) for cell in raw_rows[0]]
    rows: List[Dict[str, Any]] = []
    for source_row, row in enumerate(raw_rows[1:], start=2):
        if _is_blank_row(row):
            continue
        item: Dict[str, Any] = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        item["_source_sheet"] = sheet_name
        item["_source_row"] = source_row
        rows.append(item)
    return headers, rows


def _load_csv_manifest(manifest_bytes: bytes) -> List[Dict[str, Any]]:
    text = manifest_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows: List[Dict[str, Any]] = []
    for row_number, row in enumerate(reader, start=2):
        item = {_header_key(key): value for key, value in row.items()}
        item["_source_sheet"] = "benchmark_manifest_full.csv"
        item["_source_row"] = row_number
        rows.append(item)
    return rows


def _load_manifest_rows(
    file_bytes: bytes,
    *,
    manifest_bytes: Optional[bytes] = None,
) -> List[Dict[str, Any]]:
    if manifest_bytes:
        if manifest_bytes[:2] == b"PK":
            wb = _load_workbook(manifest_bytes)
            try:
                return _manifest_rows_from_workbook(wb)
            finally:
                wb.close()
        return _load_csv_manifest(manifest_bytes)

    wb = _load_workbook(file_bytes)
    try:
        return _manifest_rows_from_workbook(wb)
    finally:
        wb.close()


def _manifest_rows_from_workbook(wb: Any) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        headers = set(_sheet_headers(wb[sheet_name]))
        if not MANIFEST_REQUIRED_COLUMNS.issubset(headers):
            continue
        _headers, sheet_rows = _read_sheet_rows(wb, sheet_name)
        rows.extend(sheet_rows)
    return rows


def _normalize_run(value: Any) -> Optional[int]:
    denom = normalize_denomination(value)
    return denom if denom is not None else None


def _parse_numeric(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("%", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_vote_ratio(value: Any) -> Optional[float]:
    numeric = _parse_numeric(value)
    if numeric is None:
        return None
    if numeric > 1:
        return numeric / 100
    return numeric


def _consensus_reached(raw_row: Dict[str, Any]) -> Optional[bool]:
    for key in (
        "consensus_reached",
        "consensus",
        "has_consensus",
        "majority_vote",
    ):
        if key in raw_row:
            normalized = normalize_boolean(raw_row.get(key))
            if normalized is not None:
                return normalized

    vote_ratio = _normalize_vote_ratio(raw_row.get("agent_vote_pct"))
    if vote_ratio is not None:
        return vote_ratio + 1e-9 >= (2 / 3)

    valid_agent_count = normalize_denomination(raw_row.get("valid_agent_count"))
    if valid_agent_count is not None:
        return valid_agent_count >= 2

    return None


def _file_name_matches_image_id(image_id: Any, file_name: Any) -> bool:
    if not image_id or not file_name:
        return True
    image = str(image_id).strip().upper()
    name = str(file_name).strip().upper()
    if not image or not name:
        return True
    if image in name:
        return True
    match = re.fullmatch(r"([A-Z]+)(\d+)", image)
    if not match:
        return False
    number = match.group(2)
    numeric_tokens = re.findall(r"\d+", name)
    return any(token.zfill(len(number)) == number for token in numeric_tokens)


def _issue(
    *,
    code: str,
    severity: str,
    category: str,
    message: str,
    record: Optional[Dict[str, Any]] = None,
    expected: Any = None,
    actual: Any = None,
) -> Dict[str, Any]:
    record = record or {}
    return {
        "code": code,
        "severity": severity,
        "category": category,
        "dataset_id": record.get("dataset_id"),
        "image_id": record.get("image_id"),
        "model": record.get("model"),
        "run_no": record.get("run_no"),
        "file_name": record.get("file_name"),
        "source_sheet": record.get("source_sheet") or record.get("_source_sheet"),
        "source_row": record.get("source_row") or record.get("_source_row"),
        "message": message,
        "expected": expected,
        "actual": actual,
    }


def _manifest_indexes(
    manifest_rows: List[Dict[str, Any]],
) -> Tuple[Dict[Tuple[str, str], List[Dict[str, Any]]], Dict[str, str]]:
    by_dataset_image: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    dataset_by_prefix: Dict[str, set] = defaultdict(set)
    for row in manifest_rows:
        dataset_id = str(row.get("dataset_id") or "").strip()
        image_id = str(row.get("image_id") or "").strip()
        if not dataset_id or not image_id:
            continue
        by_dataset_image[(dataset_id, image_id)].append(row)
        prefix_match = re.match(r"([A-Za-z]+)", image_id)
        if prefix_match:
            dataset_by_prefix[prefix_match.group(1).upper()].add(dataset_id)
    resolved_prefixes = {
        prefix: next(iter(values))
        for prefix, values in dataset_by_prefix.items()
        if len(values) == 1
    }
    return by_dataset_image, resolved_prefixes


def _resolve_dataset_id(
    raw_row: Dict[str, Any],
    manifest_by_dataset_image: Dict[Tuple[str, str], List[Dict[str, Any]]],
    dataset_by_prefix: Dict[str, str],
) -> str:
    if raw_row.get("dataset_id"):
        return str(raw_row.get("dataset_id")).strip()
    image_id = str(raw_row.get("image_id") or "").strip()
    file_name = str(raw_row.get("file_name") or "").strip()
    for (dataset_id, manifest_image_id), rows in manifest_by_dataset_image.items():
        if image_id and manifest_image_id == image_id:
            if not file_name:
                return dataset_id
            if any(str(row.get("file_name") or "").strip() == file_name for row in rows):
                return dataset_id
            return dataset_id
    prefix_match = re.match(r"([A-Za-z]+)", image_id)
    if prefix_match:
        dataset_id = dataset_by_prefix.get(prefix_match.group(1).upper())
        if dataset_id:
            return dataset_id
    return "UNKNOWN"


def _original_value(raw_row: Dict[str, Any], key: str) -> Any:
    return raw_row.get(key)


def _parse_record(
    raw_row: Dict[str, Any],
    *,
    model_name: str,
    field_map: Dict[str, str],
    manifest_by_dataset_image: Dict[Tuple[str, str], List[Dict[str, Any]]],
    dataset_by_prefix: Dict[str, str],
) -> Dict[str, Any]:
    ground_truth_country = raw_row.get(field_map["ground_truth_country"])
    predicted_country = raw_row.get(field_map["predicted_country"])
    ground_truth_currency = raw_row.get(field_map["ground_truth_currency"])
    predicted_currency = raw_row.get(field_map["predicted_currency"])
    ground_truth_denomination = raw_row.get(field_map["ground_truth_denomination"])
    predicted_denomination = raw_row.get(field_map["predicted_denomination"])
    correctness = calculate_field_correctness(
        ground_truth_country=ground_truth_country,
        predicted_country=predicted_country,
        ground_truth_currency=ground_truth_currency,
        predicted_currency=predicted_currency,
        ground_truth_denomination=ground_truth_denomination,
        predicted_denomination=predicted_denomination,
    )

    record = {
        "dataset_id": _resolve_dataset_id(
            raw_row,
            manifest_by_dataset_image,
            dataset_by_prefix,
        ),
        "image_id": str(raw_row.get("image_id") or "").strip() or None,
        "run_no": _normalize_run(raw_row.get("run_no") or raw_row.get("run")),
        "file_name": raw_row.get("file_name"),
        "angle": raw_row.get("angle"),
        "model": model_name,
        "source_sheet": raw_row.get("_source_sheet"),
        "source_row": raw_row.get("_source_row"),
        "ground_truth_country": ground_truth_country,
        "ground_truth_currency": ground_truth_currency,
        "ground_truth_denomination": ground_truth_denomination,
        "predicted_country": predicted_country,
        "predicted_currency": predicted_currency,
        "predicted_denomination": predicted_denomination,
        "country_correct_original": normalize_boolean(
            _original_value(raw_row, "country_correct")
        ),
        "currency_correct_original": normalize_boolean(
            _original_value(raw_row, "currency_correct")
        ),
        "denomination_correct_original": normalize_boolean(
            _original_value(raw_row, "denomination_correct")
        ),
        "field_correct_count_original": normalize_denomination(
            _original_value(raw_row, "field_correct_count")
        ),
        "field_score_pct_original": raw_row.get("field_score_pct"),
        "exact_match_original": normalize_boolean(_original_value(raw_row, "exact_match")),
        "valid_agent_count_original": normalize_denomination(
            _original_value(raw_row, "valid_agent_count")
        ),
        "agent_vote_pct_original": _parse_numeric(_original_value(raw_row, "agent_vote_pct")),
        "consensus_reached": _consensus_reached(raw_row),
        "_raw_row": raw_row,
    }
    record.update(correctness)
    record["country_correct_verified"] = correctness["country_correct"]
    record["currency_correct_verified"] = correctness["currency_correct"]
    record["denomination_correct_verified"] = correctness["denomination_correct"]
    record["field_correct_count_verified"] = correctness["field_correct_count"]
    record["field_score_pct_verified"] = correctness["field_score_pct"]
    record["exact_match_verified"] = correctness["exact_match"]
    record["exact_match_mismatch"] = (
        record["exact_match_original"] is not None
        and record["exact_match_original"] != record["exact_match_verified"]
    )
    record["country_correct_mismatch"] = (
        record["country_correct_original"] is not None
        and record["country_correct_original"]
        != record["country_correct_verified"]
    )
    record["currency_correct_mismatch"] = (
        record["currency_correct_original"] is not None
        and record["currency_correct_original"]
        != record["currency_correct_verified"]
    )
    record["denomination_correct_mismatch"] = (
        record["denomination_correct_original"] is not None
        and record["denomination_correct_original"]
        != record["denomination_correct_verified"]
    )
    record["field_correct_count_mismatch"] = (
        record["field_correct_count_original"] is not None
        and record["field_correct_count_original"]
        != record["field_correct_count_verified"]
    )
    return record


def _parse_workbook_records(
    file_bytes: bytes,
    *,
    manifest_rows: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[str], List[str]]:
    wb = _load_workbook(file_bytes)
    try:
        manifest_by_dataset_image, dataset_by_prefix = _manifest_indexes(manifest_rows)
        records: List[Dict[str, Any]] = []
        hethong_headers: List[str] = []
        gpt_headers: List[str] = []

        hethong_sheet = _find_sheet(
            wb,
            required_columns=HE_THONG_REQUIRED_COLUMNS,
            candidates=HE_THONG_SHEET_CANDIDATES,
        )
        if not hethong_sheet:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Không tìm thấy sheet hệ thống có đủ cột ground_truth_* "
                    "và predicted_*."
                ),
            )
        hethong_headers, hethong_rows = _read_sheet_rows(wb, hethong_sheet)
        ht_field_map = {
            "ground_truth_country": "ground_truth_country",
            "ground_truth_currency": "ground_truth_currency",
            "ground_truth_denomination": "ground_truth_denomination",
            "predicted_country": "predicted_country",
            "predicted_currency": "predicted_currency",
            "predicted_denomination": "predicted_denomination",
        }
        for raw_row in hethong_rows:
            records.append(
                _parse_record(
                    raw_row,
                    model_name=MODEL_BANKNOTEAI,
                    field_map=ht_field_map,
                    manifest_by_dataset_image=manifest_by_dataset_image,
                    dataset_by_prefix=dataset_by_prefix,
                )
            )

        gpt_sheet = _find_sheet(
            wb,
            required_columns=GPT_REQUIRED_COLUMNS,
            candidates=GPT_GEMINI_SHEET_CANDIDATES,
        )
        if gpt_sheet:
            gpt_headers, gpt_rows = _read_sheet_rows(wb, gpt_sheet)
            gg_field_map = {
                "ground_truth_country": "gt_country",
                "ground_truth_currency": "gt_currency",
                "ground_truth_denomination": "gt_denomination",
                "predicted_country": "pred_country",
                "predicted_currency": "pred_currency",
                "predicted_denomination": "pred_denomination",
            }
            for raw_row in gpt_rows:
                model_name = str(raw_row.get("model_name") or "").strip()
                records.append(
                    _parse_record(
                        raw_row,
                        model_name=model_name,
                        field_map=gg_field_map,
                        manifest_by_dataset_image=manifest_by_dataset_image,
                        dataset_by_prefix=dataset_by_prefix,
                    )
                )
        return records, hethong_headers, gpt_headers
    finally:
        wb.close()


def _manifest_record(row: Dict[str, Any]) -> Dict[str, Any]:
    correctness = calculate_field_correctness(
        ground_truth_country=row.get("ground_truth_country"),
        predicted_country=row.get("ground_truth_country"),
        ground_truth_currency=row.get("ground_truth_currency"),
        predicted_currency=row.get("ground_truth_currency"),
        ground_truth_denomination=row.get("ground_truth_denomination"),
        predicted_denomination=row.get("ground_truth_denomination"),
    )
    return {
        "dataset_id": str(row.get("dataset_id") or "").strip() or None,
        "image_id": str(row.get("image_id") or "").strip() or None,
        "run_no": _normalize_run(row.get("run_no") or row.get("run")),
        "file_name": row.get("file_name"),
        "model": "MASTER",
        "source_sheet": row.get("_source_sheet"),
        "source_row": row.get("_source_row"),
        "ground_truth_country": row.get("ground_truth_country"),
        "ground_truth_currency": row.get("ground_truth_currency"),
        "ground_truth_denomination": row.get("ground_truth_denomination"),
        "normalized_ground_truth_country": correctness[
            "normalized_ground_truth_country"
        ],
        "normalized_ground_truth_currency": correctness[
            "normalized_ground_truth_currency"
        ],
        "normalized_ground_truth_denomination": correctness[
            "normalized_ground_truth_denomination"
        ],
    }


def _sequence_missing_ids(manifest_records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    groups: Dict[Tuple[str, str, int], set] = defaultdict(set)
    for record in manifest_records:
        image_id = str(record.get("image_id") or "").strip()
        dataset_id = str(record.get("dataset_id") or "").strip()
        match = re.fullmatch(r"([A-Za-z]+)(\d+)", image_id)
        if not dataset_id or not match:
            continue
        groups[(dataset_id, match.group(1).upper(), len(match.group(2)))].add(
            int(match.group(2))
        )
    for (dataset_id, prefix, width), numbers in groups.items():
        if len(numbers) < 2:
            continue
        for number in range(min(numbers), max(numbers) + 1):
            if number in numbers:
                continue
            image_id = f"{prefix}{number:0{width}d}"
            issues.append(
                _issue(
                    code="MISSING_MASTER_SEQUENCE_IMAGE_ID",
                    severity="ERROR",
                    category="Missing_Image_IDs",
                    message="Image ID is missing from the benchmark master sequence.",
                    record={"dataset_id": dataset_id, "image_id": image_id},
                    expected=image_id,
                    actual=None,
                )
            )
    return issues


def _build_issues(
    records: List[Dict[str, Any]],
    manifest_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    record_datasets = {
        record.get("dataset_id")
        for record in records
        if record.get("dataset_id") and record.get("dataset_id") != "UNKNOWN"
    }
    all_manifest_records = [_manifest_record(row) for row in manifest_rows]
    manifest_records = [
        record
        for record in all_manifest_records
        if not record_datasets or record.get("dataset_id") in record_datasets
    ]

    if not manifest_records:
        issues.append(
            _issue(
                code="MASTER_MANIFEST_NOT_AVAILABLE",
                severity="WARNING",
                category="Data_Integrity",
                message="No embedded or external benchmark master manifest was available.",
            )
        )

    for record in records:
        if not record.get("image_id"):
            issues.append(
                _issue(
                    code="MISSING_IMAGE_ID",
                    severity="ERROR",
                    category="Missing_Image_IDs",
                    message="Result row is missing image_id.",
                    record=record,
                )
            )
        if not record.get("model"):
            issues.append(
                _issue(
                    code="MISSING_MODEL_NAME",
                    severity="ERROR",
                    category="Data_Integrity",
                    message="Result row is missing model_name.",
                    record=record,
                )
            )
        run_no = record.get("run_no")
        if run_no is None or run_no < 1 or run_no > 3:
            issues.append(
                _issue(
                    code="INVALID_RUN",
                    severity="ERROR",
                    category="Data_Integrity",
                    message="Run must be present and between 1 and 3.",
                    record=record,
                    expected="1..3",
                    actual=run_no,
                )
            )
        if not _file_name_matches_image_id(record.get("image_id"), record.get("file_name")):
            issues.append(
                _issue(
                    code="FILE_NAME_ID_MISMATCH",
                    severity="ERROR",
                    category="File_Name_ID_Mismatches",
                    message="file_name does not match image_id.",
                    record=record,
                    expected=record.get("image_id"),
                    actual=record.get("file_name"),
                )
            )
        if record.get("exact_match_mismatch"):
            issues.append(
                _issue(
                    code="EXACT_MATCH_MISMATCH",
                    severity="ERROR",
                    category="Data_Integrity",
                    message="Original exact_match differs from verified exact_match.",
                    record=record,
                    expected=record.get("exact_match_verified"),
                    actual=record.get("exact_match_original"),
                )
            )
        if not record.get("has_complete_prediction"):
            missing_fields = [
                field
                for field in ("country", "currency", "denomination")
                if record.get(f"normalized_predicted_{field}") is None
            ]
            issues.append(
                _issue(
                    code="MISSING_PREDICTION",
                    severity="WARNING",
                    category="Data_Integrity",
                    message=(
                        "Prediction is missing one or more required fields: "
                        + ", ".join(missing_fields)
                    ),
                    record=record,
                    expected="country|currency|denomination",
                    actual=_composite_label(
                        record,
                        (
                            "normalized_predicted_country",
                            "normalized_predicted_currency",
                            "normalized_predicted_denomination",
                        ),
                    ),
                )
            )
        for field in ("country", "currency", "denomination"):
            if record.get(f"{field}_correct_mismatch"):
                issues.append(
                    _issue(
                        code=f"{field.upper()}_CORRECT_MISMATCH",
                        severity="WARNING",
                        category="Data_Integrity",
                        message=(
                            f"Original {field}_correct differs from verified "
                            f"{field}_correct."
                        ),
                        record=record,
                        expected=record.get(f"{field}_correct_verified"),
                        actual=record.get(f"{field}_correct_original"),
                    )
                )
        if record.get("field_correct_count_mismatch"):
            issues.append(
                _issue(
                    code="FIELD_CORRECT_COUNT_MISMATCH",
                    severity="ERROR",
                    category="Data_Integrity",
                    message="Original field_correct_count differs from verified count.",
                    record=record,
                    expected=record.get("field_correct_count_verified"),
                    actual=record.get("field_correct_count_original"),
                )
            )

    duplicate_groups: Dict[Tuple[Any, Any, Any, Any], List[Dict[str, Any]]] = defaultdict(list)
    for record in records:
        key = (
            record.get("dataset_id"),
            record.get("image_id"),
            record.get("model"),
            record.get("run_no"),
        )
        duplicate_groups[key].append(record)
    for key, items in duplicate_groups.items():
        if any(value in (None, "") for value in key):
            continue
        if len(items) <= 1:
            continue
        for record in items:
            issues.append(
                _issue(
                    code="DUPLICATE_EVAL_KEY",
                    severity="ERROR",
                    category="Duplicate_Rows",
                    message=(
                        "Duplicate dataset_id + image_id + model_name + run key."
                    ),
                    record=record,
                    expected="unique key",
                    actual=" | ".join(str(value) for value in key),
                )
            )

    master_duplicate_groups: Dict[Tuple[Any, Any], List[Dict[str, Any]]] = defaultdict(list)
    for record in manifest_records:
        key = (record.get("dataset_id"), record.get("image_id"))
        if any(value in (None, "") for value in key):
            continue
        master_duplicate_groups[key].append(record)
    for key, items in master_duplicate_groups.items():
        if len(items) <= 1:
            continue
        for record in items:
            issues.append(
                _issue(
                    code="MASTER_DUPLICATE_IMAGE_ID",
                    severity="ERROR",
                    category="Duplicate_Rows",
                    message="Duplicate image_id in benchmark master manifest.",
                    record=record,
                    expected="unique dataset_id + image_id in master",
                    actual=" | ".join(str(value) for value in key),
                )
            )

    gt_groups: Dict[Tuple[Any, Any], Dict[Tuple[Any, Any, Any], List[Dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    for record in [*records, *manifest_records]:
        dataset_id = record.get("dataset_id")
        image_id = record.get("image_id")
        if not dataset_id or not image_id:
            continue
        gt_key = (
            normalize_country(record.get("ground_truth_country")),
            normalize_currency(record.get("ground_truth_currency")),
            normalize_denomination(record.get("ground_truth_denomination")),
        )
        if all(value is None for value in gt_key):
            continue
        gt_groups[(dataset_id, image_id)][gt_key].append(record)
    for (_dataset_id, _image_id), variants in gt_groups.items():
        if len(variants) <= 1:
            continue
        for records_for_variant in variants.values():
            for record in records_for_variant:
                issues.append(
                    _issue(
                        code="GROUND_TRUTH_CONFLICT",
                        severity="ERROR",
                        category="Ground_Truth_Conflicts",
                        message=(
                            "Same dataset_id + image_id has multiple normalized "
                            "ground-truth values."
                        ),
                        record=record,
                        expected="single normalized GT tuple",
                        actual=str(list(variants.keys())),
                    )
                )

    master_ids = {
        (record.get("dataset_id"), record.get("image_id"))
        for record in manifest_records
        if record.get("dataset_id") and record.get("image_id")
    }
    result_ids = {
        (record.get("dataset_id"), record.get("image_id"))
        for record in records
        if record.get("dataset_id") and record.get("image_id")
    }
    if master_ids:
        for dataset_id, image_id in sorted(master_ids - result_ids):
            issues.append(
                _issue(
                    code="MASTER_IMAGE_ID_MISSING_IN_RESULTS",
                    severity="ERROR",
                    category="Missing_Image_IDs",
                    message="image_id exists in benchmark master but not in results.",
                    record={"dataset_id": dataset_id, "image_id": image_id},
                )
            )
        for dataset_id, image_id in sorted(result_ids - master_ids):
            issues.append(
                _issue(
                    code="UNEXPECTED_IMAGE_ID",
                    severity="ERROR",
                    category="Unexpected_Image_IDs",
                    message="image_id appears in results but not in benchmark master.",
                    record={"dataset_id": dataset_id, "image_id": image_id},
                )
            )
    issues.extend(_sequence_missing_ids(manifest_records))

    file_name_groups: Dict[Tuple[Any, Any], set] = defaultdict(set)
    for record in [*records, *manifest_records]:
        dataset_id = record.get("dataset_id")
        image_id = record.get("image_id")
        file_name = record.get("file_name")
        if dataset_id and image_id and file_name:
            file_name_groups[(dataset_id, image_id)].add(str(file_name))
    for (dataset_id, image_id), file_names in file_name_groups.items():
        if len(file_names) <= 1:
            continue
        issues.append(
            _issue(
                code="IMAGE_ID_MULTIPLE_FILE_NAMES",
                severity="ERROR",
                category="File_Name_ID_Mismatches",
                message="One image_id is associated with multiple file_name values.",
                record={"dataset_id": dataset_id, "image_id": image_id},
                expected="single file_name",
                actual=" | ".join(sorted(file_names)),
            )
        )

    issues.extend(_logic_validation_issues(records))
    return issues


def _valid_score_pct(value: Any) -> bool:
    try:
        numeric = round(float(value), 2)
    except (TypeError, ValueError):
        return False
    return any(abs(numeric - expected) <= 0.01 for expected in (0, 33.33, 66.67, 100))


def _logic_validation_issues(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    for record in records:
        exact = bool(record.get("exact_match_verified"))
        field_flags = (
            bool(record.get("country_correct_verified")),
            bool(record.get("currency_correct_verified")),
            bool(record.get("denomination_correct_verified")),
        )
        count = record.get("field_correct_count_verified")
        if exact and (not all(field_flags) or count != 3):
            issues.append(
                _issue(
                    code="EXACT_MATCH_FIELD_INCONSISTENT",
                    severity="ERROR",
                    category="Logic_Validation_Errors",
                    message=(
                        "exact_match_verified TRUE requires all three field "
                        "flags TRUE and field_correct_count_verified == 3."
                    ),
                    record=record,
                )
            )
        if count not in (0, 1, 2, 3):
            issues.append(
                _issue(
                    code="FIELD_COUNT_OUT_OF_RANGE",
                    severity="ERROR",
                    category="Logic_Validation_Errors",
                    message="field_correct_count_verified must be between 0 and 3.",
                    record=record,
                    expected="0..3",
                    actual=count,
                )
            )
        if not _valid_score_pct(record.get("field_score_pct_verified")):
            issues.append(
                _issue(
                    code="FIELD_SCORE_INVALID",
                    severity="ERROR",
                    category="Logic_Validation_Errors",
                    message="field_score_pct_verified must be 0, 33.33, 66.67 or 100.",
                    record=record,
                    actual=record.get("field_score_pct_verified"),
                )
            )

    grouped: Dict[Tuple[Any, Any], List[Dict[str, Any]]] = defaultdict(list)
    for record in records:
        grouped[(record.get("dataset_id"), record.get("model"))].append(record)
    for (_dataset_id, _model), items in grouped.items():
        total = len(items)
        complete = sum(bool(item.get("has_complete_prediction")) for item in items)
        exact = sum(bool(item.get("exact_match_verified")) for item in items)
        country = sum(bool(item.get("country_correct_verified")) for item in items)
        currency = sum(bool(item.get("currency_correct_verified")) for item in items)
        denom = sum(bool(item.get("denomination_correct_verified")) for item in items)
        first = items[0] if items else {}
        if complete > total:
            issues.append(
                _issue(
                    code="COMPLETE_PREDICTION_COUNT_INVALID",
                    severity="ERROR",
                    category="Logic_Validation_Errors",
                    message="complete_prediction_count cannot exceed total_samples.",
                    record=first,
                )
            )
        if exact > complete:
            issues.append(
                _issue(
                    code="CORRECT_COUNT_EXCEEDS_COMPLETE_PREDICTIONS",
                    severity="ERROR",
                    category="Logic_Validation_Errors",
                    message="exact_match count cannot exceed complete prediction count.",
                    record=first,
                    expected=complete,
                    actual=exact,
                )
            )
        for field_name, field_count in (
            ("country", country),
            ("currency", currency),
            ("denomination", denom),
        ):
            if exact > field_count:
                issues.append(
                    _issue(
                        code="EXACT_COUNT_EXCEEDS_FIELD_COUNT",
                        severity="ERROR",
                        category="Logic_Validation_Errors",
                        message=(
                            "exact_match_count cannot be greater than any "
                            "field-level correct count."
                        ),
                        record=first,
                        expected=f"{field_name}_correct_count >= {exact}",
                        actual=field_count,
                    )
                )
        if total and complete == 0:
            issues.append(
                _issue(
                    code="NO_COMPLETE_PREDICTIONS",
                    severity="ERROR",
                    category="Logic_Validation_Errors",
                    message="No rows have complete country/currency/denomination prediction.",
                    record=first,
                )
            )
    return issues


def _label(value: Any) -> str:
    return MISSING_CLASS_LABEL if value is None else str(value)


def _classification_summary_from_labels(
    y_true: Sequence[Any],
    y_pred: Sequence[Any],
) -> Dict[str, float]:
    if len(y_true) != len(y_pred):
        raise ValueError("y_true and y_pred must have the same length.")

    total = len(y_true)
    labels = sorted({str(value) for value in y_true} | {str(value) for value in y_pred})
    if not labels:
        return {
            "accuracy": 0.0,
            "macro_precision": 0.0,
            "macro_recall": 0.0,
            "macro_f1": 0.0,
        }

    macro_precision = 0.0
    macro_recall = 0.0
    macro_f1 = 0.0
    for label in labels:
        true_matches = [str(value) == label for value in y_true]
        pred_matches = [str(value) == label for value in y_pred]
        tp = sum(t and p for t, p in zip(true_matches, pred_matches))
        fp = sum((not t) and p for t, p in zip(true_matches, pred_matches))
        fn = sum(t and (not p) for t, p in zip(true_matches, pred_matches))

        precision = tp / (tp + fp) if (tp + fp) else 0.0
        recall = tp / (tp + fn) if (tp + fn) else 0.0
        f1 = (
            2 * precision * recall / (precision + recall)
            if (precision + recall)
            else 0.0
        )
        macro_precision += precision
        macro_recall += recall
        macro_f1 += f1

    label_count = len(labels)
    correct = sum(str(a) == str(b) for a, b in zip(y_true, y_pred))
    return {
        "accuracy": correct / total if total else 0.0,
        "macro_precision": macro_precision / label_count,
        "macro_recall": macro_recall / label_count,
        "macro_f1": macro_f1 / label_count,
    }


def _dimension_summary(
    records: List[Dict[str, Any]],
    *,
    gt_key: str,
    pred_key: str,
) -> Dict[str, float]:
    y_true = [_label(record.get(gt_key)) for record in records]
    y_pred = [_label(record.get(pred_key)) for record in records]
    return _classification_summary_from_labels(y_true, y_pred)


def _dimension_summary_valid_only(
    records: List[Dict[str, Any]],
    *,
    gt_key: str,
    pred_key: str,
) -> Dict[str, float]:
    filtered = [record for record in records if record.get(pred_key) is not None]
    y_true = [_label(record.get(gt_key)) for record in filtered]
    y_pred = [_label(record.get(pred_key)) for record in filtered]
    return _classification_summary_from_labels(y_true, y_pred)


def _composite_label(record: Dict[str, Any], keys: Sequence[str]) -> str:
    return "|".join(_label(record.get(key)) for key in keys)


def _composite_report(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    y_true = [
        _composite_label(
            record,
            (
                "normalized_ground_truth_country",
                "normalized_ground_truth_currency",
                "normalized_ground_truth_denomination",
            ),
        )
        for record in records
    ]
    y_pred = [
        _composite_label(
            record,
            (
                "normalized_predicted_country",
                "normalized_predicted_currency",
                "normalized_predicted_denomination",
            ),
        )
        for record in records
    ]
    return classification_report_dict(y_true, y_pred)


def _consensus_rate(records: List[Dict[str, Any]]) -> Optional[float]:
    consensus_values = [
        record.get("consensus_reached")
        for record in records
        if record.get("consensus_reached") is not None
    ]
    if not consensus_values:
        return None
    return sum(bool(value) for value in consensus_values) / len(records) if records else 0.0


def _dimension_report(
    records: List[Dict[str, Any]],
    *,
    gt_key: str,
    pred_key: str,
) -> Dict[str, Any]:
    y_true = [_label(record.get(gt_key)) for record in records]
    y_pred = [_label(record.get(pred_key)) for record in records]
    return classification_report_dict(y_true, y_pred)


def _dimension_report_valid_only(
    records: List[Dict[str, Any]],
    *,
    gt_key: str,
    pred_key: str,
) -> Dict[str, Any]:
    filtered = [record for record in records if record.get(pred_key) is not None]
    y_true = [_label(record.get(gt_key)) for record in filtered]
    y_pred = [_label(record.get(pred_key)) for record in filtered]
    return classification_report_dict(y_true, y_pred)


def _group_issues(
    issues: List[Dict[str, Any]],
    *,
    dataset_id: Any,
    model: Any,
) -> List[Dict[str, Any]]:
    result = []
    for issue in issues:
        issue_dataset = issue.get("dataset_id")
        issue_model = issue.get("model")
        if issue_dataset not in (None, "", dataset_id):
            continue
        if issue_model not in (None, "", model, "MASTER"):
            continue
        result.append(issue)
    return result


def _metric_status(group_records: List[Dict[str, Any]], group_issues: List[Dict[str, Any]]) -> str:
    serious_codes = {
        "DUPLICATE_EVAL_KEY",
        "EXACT_MATCH_MISMATCH",
        "GROUND_TRUTH_CONFLICT",
        "NO_COMPLETE_PREDICTIONS",
    }
    if not group_records:
        return "INVALID"
    if any(
        issue.get("severity") == "ERROR"
        and issue.get("code") in serious_codes
        for issue in group_issues
    ):
        return "INVALID"
    if any(issue.get("severity") == "ERROR" for issue in group_issues):
        return "PROVISIONAL"
    if any(issue.get("severity") == "WARNING" for issue in group_issues):
        return "PROVISIONAL"
    return "VALID"


def _data_integrity_status(group_records: List[Dict[str, Any]], group_issues: List[Dict[str, Any]]) -> str:
    has_original_exact = any(
        record.get("exact_match_original") is not None for record in group_records
    )
    if any(issue.get("severity") == "ERROR" for issue in group_issues):
        return "FAILED"
    if not has_original_exact:
        return "NOT_AVAILABLE"
    mismatch_count = sum(bool(record.get("exact_match_mismatch")) for record in group_records)
    return "FAILED" if mismatch_count else "PASS"


def _compute_group_metrics(
    records: List[Dict[str, Any]],
    issues: List[Dict[str, Any]],
) -> Dict[str, Any]:
    total = len(records)
    exact_verified_count = sum(bool(row.get("exact_match_verified")) for row in records)
    complete_count = sum(bool(row.get("has_complete_prediction")) for row in records)
    missing_count = total - complete_count
    exact_complete = sum(
        bool(row.get("exact_match_verified")) and bool(row.get("has_complete_prediction"))
        for row in records
    )
    has_original_exact = any(row.get("exact_match_original") is not None for row in records)
    exact_original_count = (
        sum(bool(row.get("exact_match_original")) for row in records)
        if has_original_exact
        else None
    )
    mismatch_count = (
        sum(bool(row.get("exact_match_mismatch")) for row in records)
        if has_original_exact
        else None
    )

    dimensions = {
        "Country": (
            "normalized_ground_truth_country",
            "normalized_predicted_country",
            "country_correct_verified",
        ),
        "Currency": (
            "normalized_ground_truth_currency",
            "normalized_predicted_currency",
            "currency_correct_verified",
        ),
        "Denomination": (
            "normalized_ground_truth_denomination",
            "normalized_predicted_denomination",
            "denomination_correct_verified",
        ),
    }
    dimension_metrics: Dict[str, Dict[str, Any]] = {}
    for dimension, (gt_key, pred_key, correct_key) in dimensions.items():
        report = _dimension_report(records, gt_key=gt_key, pred_key=pred_key)
        valid_report = _dimension_report_valid_only(
            records,
            gt_key=gt_key,
            pred_key=pred_key,
        )
        summary = _dimension_summary(records, gt_key=gt_key, pred_key=pred_key)
        valid_summary = _dimension_summary_valid_only(
            records,
            gt_key=gt_key,
            pred_key=pred_key,
        )
        valid_prediction_samples = sum(row.get(pred_key) is not None for row in records)
        dimension_metrics[dimension] = {
            "report": report,
            "valid_report": valid_report,
            "summary": summary,
            "valid_summary": valid_summary,
            "valid_prediction_samples": valid_prediction_samples,
            "missing_prediction_samples": total - valid_prediction_samples,
            "accuracy": (
                sum(bool(row.get(correct_key)) for row in records) / total
                if total
                else 0.0
            ),
        }

    average_dimension_macro_precision = mean(
        metric["summary"]["macro_precision"] for metric in dimension_metrics.values()
    )
    average_dimension_macro_recall = mean(
        metric["summary"]["macro_recall"] for metric in dimension_metrics.values()
    )
    average_dimension_macro_f1 = mean(
        metric["summary"]["macro_f1"] for metric in dimension_metrics.values()
    )
    consensus_rate = _consensus_rate(records)

    group_issues = _group_issues(
        issues,
        dataset_id=records[0].get("dataset_id") if records else None,
        model=records[0].get("model") if records else None,
    )
    return {
        "total_samples": total,
        "unique_images": len({row.get("image_id") for row in records if row.get("image_id")}),
        "complete_prediction_count": complete_count,
        "missing_prediction_count": missing_count,
        "coverage": complete_count / total if total else 0.0,
        "correct_samples": exact_verified_count,
        "accuracy": exact_verified_count / total if total else 0.0,
        "macro_precision": average_dimension_macro_precision,
        "macro_recall": average_dimension_macro_recall,
        "macro_f1": average_dimension_macro_f1,
        "consensus_rate": consensus_rate,
        "exact_match_verified_count": exact_verified_count,
        "end_to_end_accuracy": exact_verified_count / total if total else 0.0,
        "conditional_accuracy": (
            exact_complete / complete_count if complete_count else None
        ),
        "accuracy_official": (
            exact_original_count / total if total and exact_original_count is not None else None
        ),
        "accuracy_verification": exact_verified_count / total if total else 0.0,
        "exact_match_mismatch_count": mismatch_count,
        "country_accuracy": dimension_metrics["Country"]["accuracy"],
        "currency_accuracy": dimension_metrics["Currency"]["accuracy"],
        "denomination_accuracy": dimension_metrics["Denomination"]["accuracy"],
        "average_dimension_macro_precision": average_dimension_macro_precision,
        "average_dimension_macro_recall": average_dimension_macro_recall,
        "average_dimension_macro_f1": average_dimension_macro_f1,
        "dimension_metrics": dimension_metrics,
        "data_integrity_status": _data_integrity_status(records, group_issues),
        "metric_status": _metric_status(records, group_issues),
    }


def _group_records(records: List[Dict[str, Any]], key_fields: Sequence[str]) -> Dict[Tuple[Any, ...], List[Dict[str, Any]]]:
    groups: Dict[Tuple[Any, ...], List[Dict[str, Any]]] = defaultdict(list)
    for record in records:
        key = tuple(record.get(field) for field in key_fields)
        groups[key].append(record)
    return groups


def _raw_rows_for_model(
    records: List[Dict[str, Any]],
    *,
    headers: List[str],
    model_name: str,
) -> Tuple[List[str], List[List[Any]]]:
    selected = [record for record in records if record.get("model") == model_name]
    all_headers = [*headers, *RAW_VERIFIED_COLUMNS]
    rows: List[List[Any]] = []
    for record in selected:
        raw_row = record.get("_raw_row") or {}
        rows.append(
            [raw_row.get(header) for header in headers]
            + [
                record.get("dataset_id"),
                record.get("model"),
                record.get("source_sheet"),
                record.get("source_row"),
                record.get("normalized_ground_truth_country"),
                record.get("normalized_predicted_country"),
                record.get("normalized_ground_truth_currency"),
                record.get("normalized_predicted_currency"),
                record.get("normalized_ground_truth_denomination"),
                record.get("normalized_predicted_denomination"),
                record.get("country_correct_original"),
                record.get("currency_correct_original"),
                record.get("denomination_correct_original"),
                record.get("field_correct_count_original"),
                record.get("field_score_pct_original"),
                record.get("exact_match_original"),
                record.get("country_correct_verified"),
                record.get("currency_correct_verified"),
                record.get("denomination_correct_verified"),
                record.get("field_correct_count_verified"),
                record.get("field_score_pct_verified"),
                record.get("exact_match_verified"),
                record.get("has_complete_prediction"),
                record.get("exact_match_mismatch"),
                record.get("country_correct_mismatch"),
                record.get("currency_correct_mismatch"),
                record.get("denomination_correct_mismatch"),
                record.get("field_correct_count_mismatch"),
            ]
        )
    return all_headers, rows


def _summary_rows(records: List[Dict[str, Any]], issues: List[Dict[str, Any]]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for (dataset_id, model), items in sorted(
        _group_records(records, ("dataset_id", "model")).items(),
        key=lambda item: str(item[0]),
    ):
        metrics = _compute_group_metrics(items, issues)
        rows.append(
            [
                dataset_id,
                model,
                metrics["total_samples"],
                metrics["complete_prediction_count"],
                metrics["correct_samples"],
                metrics["coverage"],
                metrics["accuracy"],
                metrics["macro_precision"],
                metrics["macro_recall"],
                metrics["macro_f1"],
                metrics["consensus_rate"],
                metrics["metric_status"],
            ]
        )
    return rows


def _dimension_rows(records: List[Dict[str, Any]], issues: List[Dict[str, Any]]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for (dataset_id, model), items in sorted(
        _group_records(records, ("dataset_id", "model")).items(),
        key=lambda item: str(item[0]),
    ):
        metrics = _compute_group_metrics(items, issues)
        for dimension, metric in metrics["dimension_metrics"].items():
            summary = metric["summary"]
            rows.append(
                [
                    dataset_id,
                    model,
                    dimension,
                    summary["accuracy"],
                    summary["macro_precision"],
                    summary["macro_recall"],
                    summary["macro_f1"],
                ]
            )
    return rows


def _per_run_rows(records: List[Dict[str, Any]], issues: List[Dict[str, Any]]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for (dataset_id, model, run_no), items in sorted(
        _group_records(records, ("dataset_id", "model", "run_no")).items(),
        key=lambda item: str(item[0]),
    ):
        metrics = _compute_group_metrics(items, issues)
        rows.append(
            [
                dataset_id,
                model,
                run_no,
                metrics["total_samples"],
                metrics["coverage"],
                metrics["accuracy"],
                metrics["macro_precision"],
                metrics["macro_recall"],
                metrics["macro_f1"],
                metrics["consensus_rate"],
            ]
        )
    return rows


def _run_stability_rows(per_run_rows: List[List[Any]]) -> Tuple[List[str], List[List[Any]]]:
    headers = ["Dataset", "Model", "Metric", "Mean", "Standard Deviation", "Min", "Max"]
    index = {name: i for i, name in enumerate(PER_RUN_HEADERS)}
    numeric_metrics = [
        "Coverage",
        "End-to-End Accuracy",
        "Conditional Accuracy",
        "Country Accuracy",
        "Currency Accuracy",
        "Denomination Accuracy",
        "Average Dimension Macro Precision",
        "Average Dimension Macro Recall",
        "Average Dimension Macro F1",
    ]
    groups: Dict[Tuple[Any, Any], List[List[Any]]] = defaultdict(list)
    for row in per_run_rows:
        groups[(row[index["Dataset"]], row[index["Model"]])].append(row)

    rows: List[List[Any]] = []
    for (dataset_id, model), group_rows in sorted(groups.items(), key=lambda item: str(item[0])):
        for metric in numeric_metrics:
            values = [
                float(row[index[metric]])
                for row in group_rows
                if row[index[metric]] is not None
            ]
            if not values:
                rows.append([dataset_id, model, metric, None, None, None, None])
                continue
            rows.append(
                [
                    dataset_id,
                    model,
                    metric,
                    mean(values),
                    stdev(values) if len(values) > 1 else 0.0,
                    min(values),
                    max(values),
                ]
            )
    return headers, rows


def _issue_type(issue: Dict[str, Any]) -> str:
    code = str(issue.get("code") or "")
    if code == "GROUND_TRUTH_CONFLICT":
        return "Ground Truth Conflict"
    if code in {"DUPLICATE_EVAL_KEY", "MASTER_DUPLICATE_IMAGE_ID"}:
        return "Duplicate Row"
    if code in {
        "MISSING_IMAGE_ID",
        "MISSING_MASTER_SEQUENCE_IMAGE_ID",
        "MASTER_IMAGE_ID_MISSING_IN_RESULTS",
        "UNEXPECTED_IMAGE_ID",
    }:
        return "Missing Image ID"
    if code in {"FILE_NAME_ID_MISMATCH", "IMAGE_ID_MULTIPLE_FILE_NAMES"}:
        return "File Name Mismatch"
    if code == "EXACT_MATCH_MISMATCH":
        return "Original Exact Match Mismatch"
    if code == "MISSING_PREDICTION":
        return "Missing Prediction"
    return "Logic Validation Error"


def _issue_rows(issues: List[Dict[str, Any]], category: Optional[str] = None) -> List[List[Any]]:
    selected = [
        issue
        for issue in issues
        if category is None or issue.get("category") == category
    ]
    return [
        [
            issue.get("severity"),
            issue.get("dataset_id"),
            issue.get("model"),
            issue.get("image_id"),
            issue.get("run_no"),
            _issue_type(issue),
            issue.get("actual"),
            issue.get("expected"),
            issue.get("message"),
        ]
        for issue in selected
    ]


def _note_rows(issues: List[Dict[str, Any]]) -> List[List[Any]]:
    rows: List[List[Any]] = [
        [
            "INFO",
            None,
            None,
            None,
            None,
            "Metric Definition",
            None,
            None,
            (
                "Accuracy = Correct Samples / Total Samples. Correct Samples "
                "requires Country, Currency and Denomination all match Ground Truth."
            ),
        ],
        [
            "INFO",
            None,
            None,
            None,
            None,
            "Metric Definition",
            None,
            None,
            (
                "Precision, Recall and F1 are macro averages computed separately "
                "for Country, Currency and Denomination, then averaged across the "
                "three dimensions."
            ),
        ],
        [
            "INFO",
            None,
            None,
            None,
            None,
            "Metric Definition",
            None,
            None,
            "Consensus Rate is reported separately and is not used as Accuracy.",
        ],
    ]
    selected_codes = {
        "DUPLICATE_EVAL_KEY",
        "EXACT_MATCH_MISMATCH",
        "FILE_NAME_ID_MISMATCH",
        "GROUND_TRUTH_CONFLICT",
        "IMAGE_ID_MULTIPLE_FILE_NAMES",
        "MASTER_DUPLICATE_IMAGE_ID",
        "MASTER_IMAGE_ID_MISSING_IN_RESULTS",
        "MISSING_IMAGE_ID",
        "MISSING_MASTER_SEQUENCE_IMAGE_ID",
        "UNEXPECTED_IMAGE_ID",
    }
    noisy_codes = {
        "COUNTRY_CORRECT_MISMATCH",
        "CURRENCY_CORRECT_MISMATCH",
        "DENOMINATION_CORRECT_MISMATCH",
        "FIELD_CORRECT_COUNT_MISMATCH",
        "MISSING_PREDICTION",
    }
    seen = set()
    for issue in issues:
        code = str(issue.get("code") or "")
        if code in noisy_codes:
            continue
        if code not in selected_codes and issue.get("category") != "Logic_Validation_Errors":
            continue
        note_type = _issue_type(issue)
        key = (
            note_type,
            issue.get("dataset_id"),
            issue.get("image_id"),
            str(issue.get("expected")),
            str(issue.get("actual")),
            issue.get("message"),
        )
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            [
                issue.get("severity"),
                issue.get("dataset_id"),
                issue.get("model"),
                issue.get("image_id"),
                issue.get("run_no"),
                note_type,
                issue.get("actual"),
                issue.get("expected"),
                issue.get("message"),
            ]
        )
    return rows


def _explainability_rows() -> Tuple[List[str], List[List[Any]]]:
    headers = ["Topic", "Explanation"]
    rows = [
        [
            "End-to-End Accuracy",
            "Verified exact matches for Country, Currency and Denomination divided by total samples. Missing predictions count as incorrect.",
        ],
        [
            "Coverage",
            "Rows where predicted_country, predicted_currency and predicted_denomination all normalize to non-missing values divided by total samples.",
        ],
        [
            "Conditional Accuracy",
            "Verified exact matches among rows with complete predictions divided by complete prediction count. It is blank when no row has a complete prediction.",
        ],
        [
            "Macro Precision/Recall/F1",
            "Per-class precision, recall and F1 are computed first, then averaged without support weights. zero_division=0 semantics are used.",
        ],
        [
            "Weighted Precision/Recall/F1",
            "Per-class precision, recall and F1 averaged with ground-truth support as weight.",
        ],
        [
            "Average Dimension Macro F1",
            "Arithmetic mean of Country, Currency and Denomination end-to-end Macro F1 values. This is not a composite-label F1.",
        ],
        [
            "Accuracy Official",
            "Accuracy calculated from the original exact_match column. It is kept only for audit.",
        ],
        [
            "Accuracy Verification",
            "Accuracy recalculated from normalized Ground Truth and Prediction. This is the main accuracy source.",
        ],
        [
            "Missing prediction handling",
            f"Correctness checks use None for missing values. Classification metrics use sentinel {MISSING_SENTINEL} in the end-to-end scope and also provide a Valid-Predictions-Only scope.",
        ],
        [
            "Data Integrity Status",
            "PASS means no verified mismatch/error was detected; FAILED means the tool found data or logic errors; NOT_AVAILABLE means the source lacks original exact_match audit data.",
        ],
        [
            "Metric Status",
            "VALID has no blocking issue, PROVISIONAL has warnings/non-blocking integrity issues, INVALID has serious issues such as GT conflict, duplicate evaluation key, exact_match mismatch or no complete predictions.",
        ],
    ]
    return headers, rows


def _write_sheet(wb: Any, title: str, headers: List[str], rows: List[List[Any]]) -> None:
    ws = wb.create_sheet(title=title[:31])
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    rate_headers = {
        "Coverage",
        "Accuracy",
        "Precision",
        "Recall",
        "F1",
        "Consensus Rate",
        "Macro Precision",
        "Macro Recall",
        "Macro F1",
    }
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if isinstance(value, float):
                header = headers[col_idx - 1] if col_idx <= len(headers) else ""
                cell.number_format = "0.00%" if header in rate_headers else "0.0000"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(max(len(headers), 1))}{max(len(rows) + 1, 1)}"
    )
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(width, 10),
            45,
        )


def _build_excel(
    records: List[Dict[str, Any]],
    hethong_headers: List[str],
    gpt_headers: List[str],
    issues: List[Dict[str, Any]],
) -> io.BytesIO:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    sheets = [
        ("Metrics_Summary", SUMMARY_HEADERS, _summary_rows(records, issues)),
        ("Metrics_Per_Dimension", DIMENSION_HEADERS, _dimension_rows(records, issues)),
        ("Notes", NOTES_HEADERS, _note_rows(issues)),
    ]
    for title, headers, rows in sheets:
        _write_sheet(wb, title, headers, rows)

    wb.active = wb.sheetnames.index("Metrics_Summary")
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _raw_rows_for_gpt(
    records: List[Dict[str, Any]],
    headers: List[str],
) -> Tuple[List[str], List[List[Any]]]:
    all_headers = [*headers, *RAW_VERIFIED_COLUMNS]
    rows: List[List[Any]] = []
    for record in records:
        raw_row = record.get("_raw_row") or {}
        rows.append(
            [raw_row.get(header) for header in headers]
            + [
                record.get("dataset_id"),
                record.get("model"),
                record.get("source_sheet"),
                record.get("source_row"),
                record.get("normalized_ground_truth_country"),
                record.get("normalized_predicted_country"),
                record.get("normalized_ground_truth_currency"),
                record.get("normalized_predicted_currency"),
                record.get("normalized_ground_truth_denomination"),
                record.get("normalized_predicted_denomination"),
                record.get("country_correct_original"),
                record.get("currency_correct_original"),
                record.get("denomination_correct_original"),
                record.get("field_correct_count_original"),
                record.get("field_score_pct_original"),
                record.get("exact_match_original"),
                record.get("country_correct_verified"),
                record.get("currency_correct_verified"),
                record.get("denomination_correct_verified"),
                record.get("field_correct_count_verified"),
                record.get("field_score_pct_verified"),
                record.get("exact_match_verified"),
                record.get("has_complete_prediction"),
                record.get("exact_match_mismatch"),
                record.get("country_correct_mismatch"),
                record.get("currency_correct_mismatch"),
                record.get("denomination_correct_mismatch"),
                record.get("field_correct_count_mismatch"),
            ]
        )
    return all_headers, rows


class BenchmarkMetricsService:
    @staticmethod
    def calculate(
        file_bytes: bytes,
        manifest_bytes: Optional[bytes] = None,
    ) -> io.BytesIO:
        manifest_rows = _load_manifest_rows(
            file_bytes,
            manifest_bytes=manifest_bytes,
        )
        records, hethong_headers, gpt_headers = _parse_workbook_records(
            file_bytes,
            manifest_rows=manifest_rows,
        )
        if not records:
            raise HTTPException(status_code=400, detail="Workbook không có dòng dữ liệu.")
        issues = _build_issues(records, manifest_rows)
        return _build_excel(records, hethong_headers, gpt_headers, issues)


__all__ = [
    "BenchmarkMetricsService",
    "_build_issues",
    "_compute_group_metrics",
    "_parse_workbook_records",
]
