"""
metrics_service.py — Tính năng Metrics Calculator (ĐỘC LẬP)
============================================================
CẢNH BÁO: File này KHÔNG import và KHÔNG phụ thuộc vào experiment_service.py.
Mọi thay đổi ở đây không ảnh hưởng đến luồng chạy thực nghiệm (start_batch, _run_once).

Thư viện yêu cầu (đã thêm vào requirements.txt):
    pandas>=2.0.0
    openpyxl>=3.1.0
    scikit-learn>=1.3.0
"""

import io
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:  # pragma: no cover - exercised in the real FastAPI app.
    from fastapi import HTTPException
except ModuleNotFoundError:  # pragma: no cover - local logic tests only.
    class HTTPException(Exception):
        def __init__(self, status_code: int, detail: str):
            self.status_code = status_code
            self.detail = detail
            super().__init__(detail)

from app.utils.benchmark_metrics_math import classification_report_dict
from app.utils.benchmark_normalization import (
    MISSING_SENTINEL,
    normalize_country,
    normalize_currency,
    normalize_denomination,
)

# ---------------------------------------------------------------------------
# Hằng số
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = [
    "ground_truth_country",
    "predicted_country",
    "ground_truth_currency",
    "predicted_currency",
    "ground_truth_denomination",
    "predicted_denomination",
]

DIMENSIONS: List[Tuple[str, str, str]] = [
    ("Country", "ground_truth_country", "predicted_country"),
    ("Currency", "ground_truth_currency", "predicted_currency"),
    ("Denomination", "ground_truth_denomination", "predicted_denomination"),
]

# Màu header cho từng dimension (pastel, dễ nhìn)
DIMENSION_COLORS: Dict[str, str] = {
    "Country": "D6E4F0",        # xanh dương nhạt
    "Currency": "D5F5E3",       # xanh lá nhạt
    "Denomination": "FEF9E7",   # vàng nhạt
}

MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB


# ---------------------------------------------------------------------------
# Helpers — chuẩn hóa text
# ---------------------------------------------------------------------------

def _normalize_series(series: Any, normalizer: Any) -> Any:
    """Normalize a pandas Series with a shared benchmark normalizer.

    Missing predictions are represented by ``__MISSING__`` for end-to-end
    classification metrics; they are not coerced to 0 or an empty label.
    """

    def _norm(val: Any) -> str:
        normalized = normalizer(val)
        return MISSING_SENTINEL if normalized is None else str(normalized)

    return series.map(_norm)


# ---------------------------------------------------------------------------
# Core — Tính Classification Report
# ---------------------------------------------------------------------------

def _compute_metrics(df: Any) -> Dict[str, Dict[str, Any]]:
    """Tính classification_report cho 3 dimensions: Country, Currency, Denomination.

    Returns:
        Dict keyed by dimension name, value là output dict của sklearn
        classification_report (gồm per-class + macro avg + weighted avg).
    """
    results: Dict[str, Dict[str, Any]] = {}

    for dim_name, gt_col, pred_col in DIMENSIONS:
        normalizer = {
            "Country": normalize_country,
            "Currency": normalize_currency,
            "Denomination": normalize_denomination,
        }[dim_name]
        y_true = list(_normalize_series(df[gt_col], normalizer))
        y_pred = list(_normalize_series(df[pred_col], normalizer))
        valid_prediction_mask = [normalizer(value) is not None for value in df[pred_col]]
        valid_y_true = [
            label for label, valid in zip(y_true, valid_prediction_mask) if valid
        ]
        valid_y_pred = [
            label for label, valid in zip(y_pred, valid_prediction_mask) if valid
        ]

        report = classification_report_dict(y_true, y_pred)
        report["_valid_only"] = classification_report_dict(valid_y_true, valid_y_pred)
        report["_valid_prediction_samples"] = sum(valid_prediction_mask)
        report["_missing_prediction_samples"] = len(y_pred) - sum(valid_prediction_mask)
        results[dim_name] = report

    return results


# ---------------------------------------------------------------------------
# Styling Helpers
# ---------------------------------------------------------------------------

def _style_header_cell(
    cell: Any,
    bg_color: str = "2C3E50",
    font_color: str = "FFFFFF",
    bold: bool = True,
) -> None:
    """Áp dụng style cho ô header (nền tối, chữ trắng, in đậm)."""
    cell.font = Font(bold=bold, color=font_color, size=10)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _style_data_cell(cell: Any, alignment: str = "left") -> None:
    """Style cho ô dữ liệu thông thường."""
    cell.alignment = Alignment(horizontal=alignment, vertical="center")
    cell.font = Font(size=10)


def _auto_fit_columns(ws: Any, min_width: int = 8, max_width: int = 45) -> None:
    """Tự động căn chỉnh độ rộng cột theo nội dung dài nhất."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ---------------------------------------------------------------------------
# Sheet Writers
# ---------------------------------------------------------------------------

def _write_raw_data_sheet(ws: Any, df: Any) -> None:
    """Ghi Sheet 1 — Raw_Data: toàn bộ dữ liệu gốc từ file upload."""
    headers = list(df.columns)

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _style_header_cell(cell, bg_color="2C3E50", font_color="FFFFFF")

    # Data rows
    for row_idx, row_tuple in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_tuple, start=1):
            # Chuyển NaN/NaT thành None để openpyxl không bị lỗi
            import math
            if value is None:
                safe_val = None
            else:
                try:
                    if math.isnan(float(value)):
                        safe_val = None
                    else:
                        safe_val = value
                except (TypeError, ValueError):
                    safe_val = value
            cell = ws.cell(row=row_idx, column=col_idx, value=safe_val)
            _style_data_cell(cell, alignment="left")

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def _write_metrics_report_sheet(
    ws: Any,
    metrics: Dict[str, Dict[str, Any]],
    total_samples: int,
) -> None:
    """Ghi Sheet 2 — Metrics_Report.

    Layout:
        [Phần A] Overall Summary  — bảng tổng hợp 3 dimensions cạnh nhau
        [Phần B] Per-Class Breakdown — bảng chi tiết từng class label
    """
    current_row = 1
    SPECIAL_KEYS = {
        "accuracy",
        "macro avg",
        "weighted avg",
        "_valid_only",
        "_valid_prediction_samples",
        "_missing_prediction_samples",
    }

    # ──────────────────────────────────────────────────────────────────────
    # PHẦN A: Overall Summary
    # ──────────────────────────────────────────────────────────────────────

    # Title
    title_cell = ws.cell(row=current_row, column=1, value="OVERALL SUMMARY")
    title_cell.font = Font(bold=True, size=12, color="1A1A2E")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # Header hàng
    summary_headers = ["Metric", "Country", "Currency", "Denomination", "Ghi chú"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        _style_header_cell(cell, bg_color="2C3E50", font_color="FFFFFF")
    current_row += 1

    # Helper lấy số liệu từ report dict
    def _get_float(report: dict, key: str, sub: str) -> float:
        entry = report.get(key, {})
        if isinstance(entry, dict):
            return round(float(entry.get(sub, 0.0)), 4)
        return 0.0

    def _get_accuracy(report: dict) -> float:
        acc = report.get("accuracy", 0.0)
        try:
            return round(float(acc), 4)
        except (TypeError, ValueError):
            return 0.0

    def _get_support(report: dict) -> int:
        return int(report.get("weighted avg", {}).get("support", total_samples))

    def _get_valid_prediction_samples(report: dict) -> int:
        return int(report.get("_valid_prediction_samples", total_samples))

    def _get_missing_prediction_samples(report: dict) -> int:
        return int(report.get("_missing_prediction_samples", 0))

    summary_metrics: List[Tuple[str, Any, str]] = [
        ("Accuracy", _get_accuracy, "Tỷ lệ dự đoán đúng"),
        ("Macro Precision", lambda r: _get_float(r, "macro avg", "precision"), "Precision TB không trọng số"),
        ("Macro Recall", lambda r: _get_float(r, "macro avg", "recall"), "Recall TB không trọng số"),
        ("Macro F1-Score", lambda r: _get_float(r, "macro avg", "f1-score"), "F1 TB không trọng số"),
        ("Weighted Precision", lambda r: _get_float(r, "weighted avg", "precision"), "Precision TB theo support"),
        ("Weighted Recall", lambda r: _get_float(r, "weighted avg", "recall"), "Recall TB theo support"),
        ("Weighted F1-Score", lambda r: _get_float(r, "weighted avg", "f1-score"), "F1 TB theo support"),
        ("Total Samples", _get_support, "Số mẫu"),
        ("Valid Prediction Samples", _get_valid_prediction_samples, "Số mẫu có prediction hợp lệ"),
        ("Missing Prediction Samples", _get_missing_prediction_samples, "Số mẫu prediction bị thiếu"),
    ]

    zebra = ["F8F9FA", "FFFFFF"]
    for i, (metric_name, getter, note) in enumerate(summary_metrics):
        bg = zebra[i % 2]
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        # Cột Metric
        c = ws.cell(row=current_row, column=1, value=metric_name)
        c.font = Font(bold=True, size=10)
        c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center")

        # Cột từng dimension
        for col_idx, dim_name in enumerate(["Country", "Currency", "Denomination"], start=2):
            report = metrics.get(dim_name, {})
            val = getter(report)
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=10)
            if metric_name == "Accuracy":
                cell.number_format = "0.00%"
            elif metric_name == "Total Samples":
                cell.number_format = "0"

        # Cột Ghi chú
        c_note = ws.cell(row=current_row, column=5, value=note)
        c_note.fill = fill
        c_note.font = Font(size=9, italic=True, color="888888")
        c_note.alignment = Alignment(horizontal="left", vertical="center")

        current_row += 1

    current_row += 1  # Dòng trống ngăn cách

    # ──────────────────────────────────────────────────────────────────────
    # PHẦN B: Per-Class Breakdown
    # ──────────────────────────────────────────────────────────────────────

    # Title
    title2_cell = ws.cell(row=current_row, column=1, value="PER-CLASS BREAKDOWN")
    title2_cell.font = Font(bold=True, size=12, color="1A1A2E")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    title2_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    breakdown_headers = ["Dimension", "Class / Label", "Precision", "Recall", "F1-Score", "Support"]
    for col_idx, header in enumerate(breakdown_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        _style_header_cell(cell, bg_color="2C3E50", font_color="FFFFFF")
    current_row += 1

    # Ghi từng dimension
    for dim_name, _gt_col, _pred_col in DIMENSIONS:
        dim_color = DIMENSION_COLORS.get(dim_name, "FFFFFF")
        dim_fill = PatternFill(start_color=dim_color, end_color=dim_color, fill_type="solid")
        agg_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        report = metrics.get(dim_name, {})

        # Tách class thường và aggregate rows
        class_keys = sorted(
            [k for k in report if k not in SPECIAL_KEYS and not str(k).startswith("_")],
            key=lambda x: str(x).casefold(),
        )
        agg_keys = [k for k in ["macro avg", "weighted avg"] if k in report]
        all_keys = class_keys + agg_keys

        first_row_of_dim = True
        for class_key in all_keys:
            entry = report.get(class_key)
            if not isinstance(entry, dict):
                continue

            is_agg = class_key in SPECIAL_KEYS
            fill = agg_fill if is_agg else dim_fill

            # Cột Dimension — chỉ ghi ở row đầu, merge xuống các row sau
            dim_cell = ws.cell(
                row=current_row, column=1,
                value=dim_name if first_row_of_dim else "",
            )
            dim_cell.fill = dim_fill
            dim_cell.font = Font(bold=first_row_of_dim, size=10)
            dim_cell.alignment = Alignment(horizontal="center", vertical="center")
            first_row_of_dim = False

            # Cột Class/Label
            c_class = ws.cell(row=current_row, column=2, value=class_key)
            c_class.fill = fill
            c_class.font = Font(bold=is_agg, size=10, italic=is_agg)
            c_class.alignment = Alignment(horizontal="left", vertical="center")

            # Precision, Recall, F1-Score, Support
            for metric_key, col_idx in [("precision", 3), ("recall", 4), ("f1-score", 5), ("support", 6)]:
                raw_val = entry.get(metric_key, 0)
                if metric_key == "support":
                    val: Any = int(raw_val)
                    fmt = "0"
                else:
                    val = round(float(raw_val), 4)
                    fmt = "0.0000"
                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.fill = fill
                c.font = Font(bold=is_agg, size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = fmt

            current_row += 1

        current_row += 1  # Dòng trống giữa các dimensions

    # Freeze row header breakdown section không khả thi vì layout phức tạp
    # → chỉ freeze row 1 tổng quát
    ws.freeze_panes = "A2"
    _auto_fit_columns(ws, min_width=10, max_width=40)


# ---------------------------------------------------------------------------
# Builder — tạo workbook 2 sheets
# ---------------------------------------------------------------------------

def _build_metrics_workbook(
    df: Any,
    metrics: Dict[str, Dict[str, Any]],
) -> io.BytesIO:
    """Tạo file .xlsx 2 sheets: Raw_Data + Metrics_Report."""
    wb = openpyxl.Workbook()

    # Sheet 1: Raw_Data
    ws_raw = wb.active
    ws_raw.title = "Raw_Data"
    _write_raw_data_sheet(ws_raw, df)

    # Sheet 2: Metrics_Report
    ws_metrics = wb.create_sheet(title="Metrics_Report")
    ws_metrics.sheet_view.showGridLines = True
    _write_metrics_report_sheet(ws_metrics, metrics, total_samples=len(df))

    # Mở sang Metrics_Report khi file được mở
    wb.active = ws_metrics

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

class MetricsService:
    """Tính toán Metrics chuyên sâu từ file Excel thô.

    Tính năng ĐỘC LẬP — không import hay sửa experiment_service.py.
    Không được gọi từ bất kỳ luồng start_batch / _run_once nào.
    """

    @staticmethod
    def calculate_from_excel(file_bytes: bytes) -> io.BytesIO:
        """Đọc file Excel, tính Precision/Recall/F1, trả về workbook mới.

        Args:
            file_bytes: Nội dung raw của file .xlsx / .xls được upload.

        Returns:
            BytesIO chứa file .xlsx kết quả (2 sheets).

        Raises:
            HTTPException 400: File lỗi format, thiếu cột, hoặc không có dữ liệu.
            HTTPException 413: File vượt quá giới hạn kích thước (10MB).
        """
        import pandas as pd

        # Validate kích thước file
        if len(file_bytes) > MAX_FILE_SIZE_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File quá lớn. Giới hạn: {MAX_FILE_SIZE_BYTES // 1024 // 1024}MB.",
            )

        # Đọc Excel vào DataFrame
        try:
            df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Không thể đọc file Excel: {exc}",
            ) from exc

        # Validate cột bắt buộc
        missing_cols = [c for c in REQUIRED_COLUMNS if c not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"File thiếu các cột bắt buộc: {missing_cols}. "
                    "Cần có đủ: ground_truth_country, predicted_country, "
                    "ground_truth_currency, predicted_currency, "
                    "ground_truth_denomination, predicted_denomination."
                ),
            )

        # Validate có dữ liệu
        if len(df) == 0:
            raise HTTPException(
                status_code=400,
                detail="File Excel không có dòng dữ liệu nào (chỉ có header).",
            )

        # Tính metrics
        try:
            metrics = _compute_metrics(df)
        except Exception as exc:
            raise HTTPException(
                status_code=500,
                detail=f"Lỗi khi tính toán metrics: {exc}",
            ) from exc

        # Tạo workbook kết quả
        return _build_metrics_workbook(df, metrics)
