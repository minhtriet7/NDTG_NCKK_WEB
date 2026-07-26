from io import BytesIO

import openpyxl

from app.services.benchmark_metrics_service import (
    BenchmarkMetricsService,
    _logic_validation_issues,
)


HT_HEADERS = [
    "dataset_id",
    "image_id",
    "run_no",
    "file_name",
    "ground_truth_country",
    "ground_truth_currency",
    "ground_truth_denomination",
    "predicted_country",
    "predicted_currency",
    "predicted_denomination",
    "country_correct",
    "currency_correct",
    "field_correct_count",
    "field_score_pct",
    "exact_match",
    "valid_agent_count",
    "agent_vote_pct",
]


def _workbook_bytes(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HeThong"
    ws.append(HT_HEADERS)
    for row in rows:
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _sheet_rows(workbook_bytes, sheet_name):
    wb = openpyxl.load_workbook(BytesIO(workbook_bytes), data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    result = [dict(zip(headers, row)) for row in rows[1:]]
    wb.close()
    return result


def test_exact_match_mismatch_and_missing_prediction_are_audited():
    source = _workbook_bytes(
        [
            [
                "AI_100",
                "A026",
                1,
                "A026_AI_VN_VND_100000_TOP.jpg",
                "Vietnam",
                "VND",
                100000,
                "Vietnam",
                "VND",
                "10000 VND",
                "TRUE",
                "TRUE",
                3,
                100,
                "TRUE",
            ],
            [
                "AI_100",
                "A093",
                1,
                "A093_AI_EU_EUR_500_LEFT.jpg",
                "European Union",
                "EUR",
                500,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        ]
    )

    output = BenchmarkMetricsService.calculate(source).getvalue()
    wb = openpyxl.load_workbook(BytesIO(output), data_only=True)
    assert wb.sheetnames == [
        "Metrics_Summary",
        "Metrics_Per_Dimension",
        "Notes",
    ]
    wb.close()
    summary = _sheet_rows(output, "Metrics_Summary")[0]
    notes = _sheet_rows(output, "Notes")

    assert summary["Complete Predictions"] == 1
    assert summary["Correct Samples"] == 0
    assert summary["Coverage"] == 0.5
    assert summary["Accuracy"] == 0
    assert summary["Consensus Rate"] is None
    assert summary["Status"] == "INVALID"
    assert any(
        row["Note Type"] == "Original Exact Match Mismatch" for row in notes
    )
    assert all(row["Note Type"] != "Missing Prediction" for row in notes)


def test_duplicate_key_goes_to_duplicate_rows():
    row = [
        "AI_100",
        "A001",
        1,
        "A001_AI_VN_VND_1000_TOP.jpg",
        "Vietnam",
        "VND",
        1000,
        "Vietnam",
        "VND",
        "1000 VND",
        "TRUE",
        "TRUE",
        3,
        100,
        "TRUE",
    ]
    output = BenchmarkMetricsService.calculate(_workbook_bytes([row, row])).getvalue()
    notes = _sheet_rows(output, "Notes")

    assert any(row["Note Type"] == "Duplicate Row" for row in notes)


def test_ground_truth_conflict_from_manifest_marks_metrics_non_valid():
    source = _workbook_bytes(
        [
            [
                "REAL_100",
                "R099",
                1,
                "MIX_099_ID_IDR_100000_back_angle_right.jpg",
                "Indonesia",
                "IDR",
                100000,
                "Indonesia",
                "IDR",
                "100000 IDR",
                "TRUE",
                "TRUE",
                3,
                100,
                "TRUE",
            ]
        ]
    )
    manifest = (
        "file_name,dataset_id,image_id,ground_truth_country,ground_truth_currency,ground_truth_denomination\n"
        "MIX_099_ID_IDR_10000_back_angle_right.jpg,REAL_100,R099,Indonesia,IDR,10000\n"
        "MIX_099_ID_IDR_100000_back_angle_right.jpg,REAL_100,R099,Indonesia,IDR,100000\n"
    ).encode()

    output = BenchmarkMetricsService.calculate(source, manifest_bytes=manifest).getvalue()
    conflicts = _sheet_rows(output, "Notes")
    summary = _sheet_rows(output, "Metrics_Summary")[0]

    assert any(row["Note Type"] == "Ground Truth Conflict" for row in conflicts)
    assert summary["Status"] != "VALID"


def test_accuracy_and_consensus_are_reported_separately():
    source = _workbook_bytes(
        [
            [
                "AI_100",
                "A001",
                1,
                "A001_AI_VN_VND_1000_TOP.jpg",
                "Vietnam",
                "VND",
                1000,
                "Vietnam",
                "VND",
                "1000 VND",
                "TRUE",
                "TRUE",
                3,
                100,
                "TRUE",
                2,
                66.67,
            ],
            [
                "AI_100",
                "A002",
                1,
                "A002_AI_VN_VND_2000_TOP.jpg",
                "Vietnam",
                "VND",
                2000,
                "Vietnam",
                "VND",
                "1000 VND",
                "TRUE",
                "TRUE",
                2,
                66.67,
                "FALSE",
                2,
                66.67,
            ],
            [
                "AI_100",
                "A003",
                1,
                "A003_AI_VN_VND_5000_TOP.jpg",
                "Vietnam",
                "VND",
                5000,
                "Vietnam",
                "VND",
                "5000 VND",
                "TRUE",
                "TRUE",
                3,
                100,
                "TRUE",
                1,
                33.33,
            ],
        ]
    )

    output = BenchmarkMetricsService.calculate(source).getvalue()
    summary = _sheet_rows(output, "Metrics_Summary")[0]

    assert abs(summary["Accuracy"] - (2 / 3)) < 1e-9
    assert summary["Consensus Rate"] == 2 / 3


def test_r004_recomputes_field_correctness_from_normalized_values():
    source = _workbook_bytes(
        [
            [
                "REAL_100",
                "R004",
                1,
                "REAL_004_VN_VND_500000_front_zoom.jpg",
                "Vietnam",
                "VND",
                500000,
                "Vietnam",
                "VND",
                "100000 VND",
                "FALSE",
                "FALSE",
                0,
                0,
                "FALSE",
                0,
                0,
            ]
        ]
    )

    output = BenchmarkMetricsService.calculate(source).getvalue()
    summary = _sheet_rows(output, "Metrics_Summary")[0]
    dimensions = _sheet_rows(output, "Metrics_Per_Dimension")

    assert summary["Complete Predictions"] == 1
    assert summary["Correct Samples"] == 0
    assert summary["Accuracy"] == 0
    assert next(row for row in dimensions if row["Dimension"] == "Country")[
        "Accuracy"
    ] == 1
    assert next(row for row in dimensions if row["Dimension"] == "Currency")[
        "Accuracy"
    ] == 1
    assert next(row for row in dimensions if row["Dimension"] == "Denomination")[
        "Accuracy"
    ] == 0


def test_consistency_validation_catches_impossible_exact_match():
    issues = _logic_validation_issues(
        [
            {
                "dataset_id": "D",
                "image_id": "I001",
                "model": "BanknoteAI",
                "run_no": 1,
                "exact_match_verified": True,
                "country_correct_verified": True,
                "currency_correct_verified": True,
                "denomination_correct_verified": False,
                "field_correct_count_verified": 2,
                "field_score_pct_verified": 66.67,
                "has_complete_prediction": True,
            }
        ]
    )

    assert any(
        issue["code"] == "EXACT_MATCH_FIELD_INCONSISTENT" for issue in issues
    )
